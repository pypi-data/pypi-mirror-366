import openpyxl
from lxml import etree
from datetime import datetime
import sys
import os

def excel_to_xml(excel_path, xml_path):
    """
    Converts an Excel file to an XML file based on the specified format.

    :param excel_path: Path to the input Excel file.
    :param xml_path: Path to the output XML file.
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {excel_path}")
        return

    datasource = etree.Element("datasource")
    entities_node = etree.SubElement(datasource, "entities")
    
    sheets_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheets_data[sheet_name] = []
        current_entity_data = None

        for row in sheet.iter_rows():
            tag = row[0].value
            if tag == "DataTag":
                uri_value = row[2].value.strip() if row[2].value else ""
                class_name = row[5].value.strip() if row[5].value else ""
                
                entity_data = {
                    "type": row[1].value,
                    "uri": uri_value,
                    "title": row[3].value,
                    "className": class_name,
                    "component": sheet_name,
                    "fields": [], "titles": [], "types": [], "formats": [], "records": [], "foreign_keys": []
                }
                sheets_data[sheet_name].append(entity_data)
                current_entity_data = entity_data
            elif tag == "DataField" and current_entity_data:
                current_entity_data["fields"] = [cell.value for cell in row[1:]]
            elif tag == "DataTitle" and current_entity_data:
                current_entity_data["titles"] = [cell.value for cell in row[1:]]
            elif tag == "DataType" and current_entity_data:
                current_entity_data["types"] = [cell.value for cell in row[1:]]
            elif tag == "DataFormat" and current_entity_data:
                current_entity_data["formats"] = [cell.value for cell in row[1:]]
            elif tag == "ForeignKey" and current_entity_data:
                current_entity_data["foreign_keys"] = [cell.value for cell in row[1:]]
            elif tag == "DataRow" and current_entity_data:
                current_entity_data["records"].append([cell.value for cell in row[1:]])

    for sheet_name, entities_data in sheets_data.items():
        for entity_data in entities_data:
            process_entity(entities_node, entity_data)

    with open(xml_path, 'wb') as f:
        f.write(etree.tostring(datasource, pretty_print=True, xml_declaration=True, encoding='utf-8'))

def process_entity(entities_node, entity_data):
    """
    Processes a single entity's data and adds it to the XML tree.
    """
    entity_node = etree.SubElement(entities_node, "entity",
                                   uri=str(entity_data.get("uri", "")),
                                   title=str(entity_data.get("title", "")),
                                   type=str(entity_data.get("type", "")),
                                   className=str(entity_data.get("className", "")),
                                   component=str(entity_data.get("component", "")))

    attributes_node = etree.SubElement(entity_node, "attributes")
    original_fields = entity_data.get("fields", [])
    fields = []
    for f in original_fields:
        sf = str(f)
        if sf == '$key':
            fields.append('_key')
        else:
            fields.append(f)
    titles = entity_data.get("titles", [])
    types = entity_data.get("types", [])
    foreign_keys = entity_data.get("foreign_keys", [])

    for i, field in enumerate(fields):
        if not field:
            continue

        title = titles[i] if i < len(titles) else ""
        field_type = types[i] if i < len(types) else ""
        is_collection = "false"
        
        if isinstance(field_type, str) and field_type.startswith("List<"):
            is_collection = "true"
            field_type = field_type[5:-1]

        attribute_props = {
            "name": str(field),
            "title": str(title),
            "type": str(field_type)
        }
        if is_collection == "true":
            attribute_props["isCollection"] = is_collection
        
        if i < len(foreign_keys) and foreign_keys[i]:
            attribute_props["foreignKey"] = str(foreign_keys[i])

        etree.SubElement(attributes_node, "attribute", **attribute_props)


    formats = entity_data.get("formats", [])
    if formats:
        formats_node = etree.SubElement(entity_node, "formats")
        for i, field in enumerate(fields):
            if not field or i >= len(formats) or not formats[i]:
                continue
            etree.SubElement(formats_node, "format", name=str(field), value=str(formats[i]))

    records_node = etree.SubElement(entity_node, "records")
    
    for record_data in entity_data.get("records", []):
        record_node = etree.SubElement(records_node, "record")
        for i, field in enumerate(fields):
            if not field or i >= len(record_data):
                continue
            value = record_data[i]
            record_node.set(str(field), format_date_value(value) if value is not None else "")

def format_date_value(value):
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime("%Y-%m-%d")
        else:
            return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)

def xml_to_excel(xml_path, excel_path):
    """
    Converts an XML file to an Excel file based on the specified format.

    :param xml_path: Path to the input XML file.
    :param excel_path: Path to the output Excel file.
    """
    try:
        tree = etree.parse(xml_path)
    except (IOError, etree.XMLSyntaxError) as e:
        print(f"Error: Cannot parse XML file at {xml_path}. {e}")
        return
        
    root = tree.getroot()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    entities_by_sheet = {}
    for entity in root.findall(".//entity"):
        sheet_name = entity.get("component")
        if sheet_name not in entities_by_sheet:
            entities_by_sheet[sheet_name] = []
        entities_by_sheet[sheet_name].append(entity)

    for sheet_name, entities in entities_by_sheet.items():
        sheet = workbook.create_sheet(title=sheet_name)
        for entity in entities:
            sheet.append(["DataTag", entity.get("type"), entity.get("uri"), entity.get("title"), "实体类",entity.get("className")])

            fields = ["DataField"]
            titles = ["DataTitle"]
            types = ["DataType"]
            foreign_keys = ["ForeignKey"]
            has_foreign_key = False
            
            attributes = entity.find("attributes")
            if attributes is not None:
                for attr in attributes.findall("attribute"):
                    attr_name = attr.get("name")
                    if attr_name:
                        fields.append(attr_name)
                        titles.append(attr.get("title"))
                        attr_type = attr.get("type")
                        if attr.get("isCollection") == "true":
                            attr_type = f"List<{attr_type}>"
                        types.append(attr_type)
                        fk = attr.get("foreignKey")
                        foreign_keys.append(fk if fk is not None else "")
                        if fk is not None:
                            has_foreign_key = True
            
            sheet.append(fields)
            sheet.append(titles)
            sheet.append(types)
            if has_foreign_key:
                sheet.append(foreign_keys)

            formats_node = entity.find("formats")
            if formats_node is not None:
                formats_row = ["DataFormat"]
                format_map = {f.get("name"): f.get("value") for f in formats_node.findall("format")}
                for field in fields[1:]:
                    formats_row.append(format_map.get(field, ""))
                sheet.append(formats_row)

            records = entity.find("records")
            if records is not None:
                for record in records.findall("record"):
                    row_data = ["DataRow"]
                    for field in fields[1:]:
                        value = record.get(field)
                        row_data.append(value if value is not None else "")
                    sheet.append(row_data)
            sheet.append([])

    workbook.save(excel_path)

def main():
    if len(sys.argv) == 2:
        input_file = sys.argv[1]
        base, ext = os.path.splitext(input_file)
        if ext == '.xlsx':
            output_file = base + '.xml'
        elif ext == '.xml':
            output_file = base + '.xlsx'
        else:
            print(f"Unsupported file extension: {ext}")
            sys.exit(1)
    elif len(sys.argv) == 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    else:
        print("Usage: python converter.py <input_file> [output_file]")
        sys.exit(1)

    _, input_ext = os.path.splitext(input_file)

    if input_ext == '.xlsx':
        print(f"Converting {input_file} to {output_file}...")
        excel_to_xml(input_file, output_file)
        print("Conversion to XML complete.")
    elif input_ext == '.xml':
        print(f"Converting {input_file} to {output_file}...")
        xml_to_excel(input_file, output_file)
        print("Conversion to Excel complete.")
    else:
        print(f"Unsupported file format: {input_ext}")
        sys.exit(1)
