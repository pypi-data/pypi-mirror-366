"""
Excel exporter for regression results using openpyxl.
Creates professional publication-quality tables.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, Any
import warnings

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..core.options import OutregOptions


class ExcelExporter:
    """Export formatted tables to Excel with professional styling."""
    
    def __init__(self, options: OutregOptions):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export")
        
        self.options = options
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel cell styles."""
        
        # Fonts
        self.title_font = Font(
            name=self.options.font_name, 
            size=self.options.font_size + 2, 
            bold=True
        )
        self.header_font = Font(
            name=self.options.font_name, 
            size=self.options.font_size, 
            bold=True
        )
        self.body_font = Font(
            name=self.options.font_name, 
            size=self.options.font_size
        )
        self.note_font = Font(
            name=self.options.font_name, 
            size=self.options.font_size - 1,
            italic=True
        )
        
        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        
        # Borders
        thin_border = Side(border_style="thin", color="000000")
        self.top_border = Border(top=thin_border)
        self.bottom_border = Border(bottom=thin_border)
        self.full_border = Border(
            top=thin_border, bottom=thin_border,
            left=thin_border, right=thin_border
        )
    
    def export(self, df: pd.DataFrame, filename: str, sheet_name: str = "Results"):
        """Export DataFrame to Excel file."""
        
        filepath = Path(filename)
        
        # Handle replace/append logic
        if self.options.append and filepath.exists():
            wb = load_workbook(str(filepath))
            # Create new sheet or use existing
            if sheet_name in wb.sheetnames:
                # Find next available position
                ws = wb[sheet_name]
                start_row = ws.max_row + 3  # Leave some space
            else:
                ws = wb.create_sheet(sheet_name)
                start_row = 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            start_row = 1
        
        # Write data to worksheet
        self._write_dataframe_to_sheet(ws, df, start_row)
        
        # Apply styling
        self._apply_styling(ws, df, start_row)
        
        # Adjust column widths
        self._adjust_column_widths(ws, df)
        
        # Save workbook
        wb.save(str(filepath))
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int):
        """Write DataFrame to worksheet starting at specified row."""
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1)
                if value is not None and value != '':
                    cell.value = value
    
    def _apply_styling(self, ws, df: pd.DataFrame, start_row: int):
        """Apply professional styling to the table."""
        
        num_rows = len(df)
        num_cols = len(df.columns)
        
        # Identify different sections
        title_rows = []
        header_row = None
        data_rows = []
        stats_rows = []
        note_rows = []
        
        for i, (idx, row) in enumerate(df.iterrows()):
            row_num = start_row + i
            first_cell = str(row.iloc[0]).strip()
            
            # Check if this is a title row
            if (self.options.title and first_cell == self.options.title) or \
               (i == 0 and first_cell and all((pd.isna(row.iloc[1:]) | (row.iloc[1:] == '')))):
                title_rows.append(row_num)
            
            # Check if this is the header row
            elif first_cell == 'Variable' or first_cell in ['Model 1', 'Model 2', 'OLS', 'Fixed Effects']:
                header_row = row_num
            
            # Check if this is a statistics row
            elif first_cell in ['Observations', 'R-squared', 'F-statistic'] or \
                 (self.options.addstat and first_cell in self.options.addstat.keys()):
                stats_rows.append(row_num)
            
            # Check if this is a notes row
            elif 'p<' in first_cell or (self.options.addnote and first_cell == self.options.addnote):
                note_rows.append(row_num)
            
            # Otherwise, it's a data row
            else:
                data_rows.append(row_num)
        
        # Apply styles to different sections
        
        # Title rows
        for row_num in title_rows:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = self.title_font
                cell.alignment = self.center_align
        
        # Header row
        if header_row:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self.bottom_border
        
        # Data rows
        for row_num in data_rows:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = self.body_font
                if col == 1:  # Variable names column
                    cell.alignment = self.left_align
                else:  # Data columns
                    cell.alignment = self.center_align
        
        # Statistics rows
        for row_num in stats_rows:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = self.body_font
                if col == 1:
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.center_align
                if row_num == stats_rows[0]:  # First stats row
                    cell.border = self.top_border
        
        # Note rows
        for row_num in note_rows:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = self.note_font
                cell.alignment = self.left_align
    
    def _adjust_column_widths(self, ws, df: pd.DataFrame):
        """Adjust column widths based on content."""
        
        for col_idx, column in enumerate(df.columns):
            col_letter = get_column_letter(col_idx + 1)
            
            # Calculate max width needed
            max_width = len(str(column))  # Header width
            
            for value in df.iloc[:, col_idx]:
                if pd.notna(value):
                    max_width = max(max_width, len(str(value)))
            
            # Set reasonable bounds
            width = min(max(max_width + 2, 8), 20)
            ws.column_dimensions[col_letter].width = width


def export_to_excel(df: pd.DataFrame, filename: str, options: OutregOptions, 
                   sheet_name: str = "Results"):
    """Convenience function for Excel export."""
    exporter = ExcelExporter(options)
    exporter.export(df, filename, sheet_name)
