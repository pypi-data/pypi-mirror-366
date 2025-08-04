"""Excel file reader implementation."""

import csv
import io
from pathlib import Path
from typing import List, Optional, Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment

from vector_db_query.document_processor.office_readers import SpreadsheetReader
from vector_db_query.document_processor.exceptions import DocumentReadError
from vector_db_query.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelReader(SpreadsheetReader):
    """Reader for Excel files (.xlsx, .xls, .csv)."""
    
    def can_read(self, file_path: Path) -> bool:
        """Check if this reader can handle the file type."""
        return file_path.suffix.lower() in self.supported_extensions
        
    def read(self, file_path: Path) -> str:
        """Read Excel file and extract text content."""
        extension = file_path.suffix.lower()
        
        try:
            if extension == '.csv':
                return self._read_csv(file_path)
            elif extension in ['.xlsx', '.xls']:
                return self._read_excel(file_path)
            else:
                raise DocumentReadError(
                    f"Unsupported Excel format: {extension}",
                    file_path=str(file_path)
                )
        except Exception as e:
            raise DocumentReadError(
                f"Failed to read Excel file: {e}",
                file_path=str(file_path)
            )
            
    @property
    def supported_extensions(self) -> List[str]:
        """Get list of supported file extensions."""
        return ['.xlsx', '.xls', '.csv']
        
    def _read_csv(self, file_path: Path) -> str:
        """Read CSV file with automatic delimiter detection."""
        logger.info(f"Reading CSV file: {file_path.name}")
        
        # First, detect the delimiter
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            sample = f.read(4096)
            sniffer = csv.Sniffer()
            try:
                delimiter = sniffer.sniff(sample).delimiter
            except csv.Error:
                delimiter = ','  # Default to comma
                
        # Read the CSV file
        try:
            df = pd.read_csv(file_path, delimiter=delimiter, dtype=str)
            
            # Extract metadata
            self._metadata = self._extract_metadata(file_path)
            self._metadata.update({
                'rows': len(df),
                'columns': len(df.columns),
                'delimiter': delimiter
            })
            
            # Convert to text
            text_parts = []
            
            # Add column headers
            headers = ' | '.join(df.columns)
            text_parts.append(headers)
            text_parts.append('-' * len(headers))
            
            # Add rows
            for _, row in df.iterrows():
                row_text = ' | '.join(str(val) if pd.notna(val) else '' for val in row)
                text_parts.append(row_text)
                
            return '\n'.join(text_parts)
            
        except Exception as e:
            logger.error(f"Error reading CSV: {e}")
            raise
            
    def _read_excel(self, file_path: Path) -> str:
        """Read Excel file (.xlsx or .xls)."""
        logger.info(f"Reading Excel file: {file_path.name}")
        
        text_parts = []
        self._metadata = self._extract_metadata(file_path)
        
        # Try openpyxl for .xlsx files
        if file_path.suffix.lower() == '.xlsx':
            text_parts = self._read_xlsx_openpyxl(file_path)
        else:
            # Use pandas for .xls files
            text_parts = self._read_xls_pandas(file_path)
            
        return self.sheet_separator.join(text_parts)
        
    def _read_xlsx_openpyxl(self, file_path: Path) -> List[str]:
        """Read .xlsx file using openpyxl for better formula and comment support."""
        text_parts = []
        
        try:
            wb = load_workbook(file_path, data_only=not self.include_formulas)
            
            self._metadata['sheets'] = wb.sheetnames
            self._metadata['total_sheets'] = len(wb.sheetnames)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Skip empty sheets
                if ws.max_row == 0 or ws.max_column == 0:
                    continue
                    
                sheet_parts = [f"Sheet: {sheet_name}", "=" * 40]
                
                # Extract data with formulas/values
                rows_data = []
                comments_data = []
                
                for row in ws.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            if self.include_formulas and hasattr(cell, 'formula') and cell.formula:
                                row_values.append(f"{cell.value} [={cell.formula}]")
                            else:
                                row_values.append(str(cell.value))
                        else:
                            row_values.append('')
                            
                        # Collect comments
                        if self.include_comments and cell.comment:
                            comments_data.append(
                                f"Cell {cell.coordinate}: {cell.comment.text}"
                            )
                            
                    rows_data.append(row_values)
                    
                # Format the data
                if rows_data:
                    formatted_text = self._format_table_data(rows_data)
                    sheet_parts.append(formatted_text)
                    
                # Add comments section
                if comments_data:
                    sheet_parts.append("\nComments:")
                    sheet_parts.extend(comments_data)
                    
                text_parts.append('\n'.join(sheet_parts))
                
        except Exception as e:
            logger.error(f"Error reading with openpyxl: {e}")
            # Fall back to pandas
            return self._read_xls_pandas(file_path)
            
        return text_parts
        
    def _read_xls_pandas(self, file_path: Path) -> List[str]:
        """Read Excel file using pandas (works for both .xls and .xlsx)."""
        text_parts = []
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            
            self._metadata['sheets'] = excel_file.sheet_names
            self._metadata['total_sheets'] = len(excel_file.sheet_names)
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
                
                # Skip empty sheets
                if df.empty:
                    continue
                    
                sheet_parts = [f"Sheet: {sheet_name}", "=" * 40]
                
                # Convert to text
                headers = ' | '.join(df.columns)
                sheet_parts.append(headers)
                sheet_parts.append('-' * len(headers))
                
                for _, row in df.iterrows():
                    row_text = ' | '.join(
                        str(val) if pd.notna(val) else '' for val in row
                    )
                    sheet_parts.append(row_text)
                    
                text_parts.append('\n'.join(sheet_parts))
                
        except Exception as e:
            logger.error(f"Error reading with pandas: {e}")
            raise
            
        return text_parts
        
    def _process_content(self, content: Any) -> str:
        """Process the raw content into text.
        
        This is implemented through the read methods above.
        """
        # Not used in this implementation
        pass