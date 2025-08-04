"""Create sample Excel files for testing."""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment

# Get the directory
sample_dir = Path(__file__).parent

# Create simple XLSX
wb = Workbook()
ws = wb.active
ws.title = "Sales Data"
ws.append(['Product', 'Q1', 'Q2', 'Q3', 'Q4', 'Total'])
ws.append(['Widget A', 100, 150, 200, 175, '=SUM(B2:E2)'])
ws.append(['Widget B', 80, 90, 110, 120, '=SUM(B3:E3)'])
ws.append(['Widget C', 200, 180, 220, 240, '=SUM(B4:E4)'])
wb.save(sample_dir / 'sales_data.xlsx')

# Create multi-sheet XLSX
wb2 = Workbook()
ws1 = wb2.active
ws1.title = "Revenue"
ws1.append(['Month', 'Revenue', 'Expenses', 'Profit'])
ws1.append(['January', 10000, 7000, '=B2-C2'])
ws1.append(['February', 12000, 8000, '=B3-C3'])

ws2 = wb2.create_sheet("Employees")
ws2.append(['Name', 'Department', 'Salary'])
ws2.append(['Alice', 'Sales', 50000])
ws2.append(['Bob', 'Engineering', 70000])
ws2.append(['Charlie', 'Marketing', 55000])

ws3 = wb2.create_sheet("Summary")
ws3.append(['Metric', 'Value'])
ws3.append(['Total Revenue', '=Revenue!B2+Revenue!B3'])
ws3.append(['Total Employees', 3])

wb2.save(sample_dir / 'company_data.xlsx')

# Create XLSX with comments
wb3 = Workbook()
ws = wb3.active
ws.title = "Project Status"
ws['A1'] = 'Task'
ws['B1'] = 'Status'
ws['C1'] = 'Priority'
ws['A2'] = 'Design Phase'
ws['B2'] = 'Complete'
ws['B2'].comment = Comment('Completed ahead of schedule', 'PM')
ws['C2'] = 'High'
ws['A3'] = 'Implementation'
ws['B3'] = 'In Progress'
ws['B3'].comment = Comment('75% complete, on track', 'Dev Lead')
ws['C3'] = 'High'
wb3.save(sample_dir / 'project_status.xlsx')

# Create CSV files
# Standard CSV
df1 = pd.DataFrame({
    'Name': ['Alice', 'Bob', 'Charlie', 'David'],
    'Age': [25, 30, 35, 28],
    'Department': ['Sales', 'Engineering', 'Marketing', 'HR'],
    'Salary': [50000, 70000, 55000, 45000]
})
df1.to_csv(sample_dir / 'employees.csv', index=False)

# Semicolon-delimited CSV
df2 = pd.DataFrame({
    'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor'],
    'Price': [999.99, 29.99, 79.99, 299.99],
    'Stock': [50, 200, 150, 75]
})
df2.to_csv(sample_dir / 'inventory.csv', sep=';', index=False)

# CSV with special characters
df3 = pd.DataFrame({
    'Company': ['ABC, Inc.', 'XYZ "Corp"', "O'Brien's LLC", 'Smith & Sons'],
    'Revenue': ['$1,000,000', '$500,000', '$750,000', '$2,000,000'],
    'Notes': ['Uses, commas', 'Has "quotes"', "Has 'apostrophes'", 'Special & chars']
})
df3.to_csv(sample_dir / 'companies.csv', index=False)

print("Sample Excel files created successfully!")
print(f"Location: {sample_dir}")
print("Files created:")
for file in sample_dir.glob('*.xlsx'):
    print(f"  - {file.name}")
for file in sample_dir.glob('*.csv'):
    print(f"  - {file.name}")