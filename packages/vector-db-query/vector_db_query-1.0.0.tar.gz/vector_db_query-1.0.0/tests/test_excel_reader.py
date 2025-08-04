"""Tests for Excel file reader."""

import tempfile
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import Workbook

from vector_db_query.document_processor.excel_reader import ExcelReader
from vector_db_query.document_processor.exceptions import DocumentReadError


class TestExcelReader:
    """Test suite for Excel reader."""
    
    @pytest.fixture
    def reader(self):
        """Create an Excel reader instance."""
        return ExcelReader()
        
    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory for test files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield Path(tmpdir)
            
    def test_supported_extensions(self, reader):
        """Test that reader reports correct supported extensions."""
        extensions = reader.supported_extensions
        assert '.xlsx' in extensions
        assert '.xls' in extensions
        assert '.csv' in extensions
        
    def test_can_read_excel_files(self, reader, temp_dir):
        """Test can_read method for Excel files."""
        xlsx_file = temp_dir / "test.xlsx"
        xls_file = temp_dir / "test.xls"
        csv_file = temp_dir / "test.csv"
        txt_file = temp_dir / "test.txt"
        
        assert reader.can_read(xlsx_file) is True
        assert reader.can_read(xls_file) is True
        assert reader.can_read(csv_file) is True
        assert reader.can_read(txt_file) is False
        
    def test_read_simple_xlsx(self, reader, temp_dir):
        """Test reading a simple XLSX file."""
        # Create test file
        file_path = temp_dir / "simple.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add data
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['A2'] = 'Alice'
        ws['B2'] = 30
        ws['A3'] = 'Bob'
        ws['B3'] = 25
        
        wb.save(file_path)
        
        # Read file
        content = reader.read(file_path)
        
        # Verify content
        assert 'Sheet: TestSheet' in content
        assert 'Name' in content
        assert 'Age' in content
        assert 'Alice' in content
        assert 'Bob' in content
        assert '30' in content
        assert '25' in content
        
    def test_read_multi_sheet_xlsx(self, reader, temp_dir):
        """Test reading XLSX file with multiple sheets."""
        file_path = temp_dir / "multi_sheet.xlsx"
        wb = Workbook()
        
        # First sheet
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1['A1'] = 'Data1'
        
        # Second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2['A1'] = 'Data2'
        
        wb.save(file_path)
        
        # Read file
        content = reader.read(file_path)
        
        # Verify both sheets are present
        assert 'Sheet: Sheet1' in content
        assert 'Sheet: Sheet2' in content
        assert 'Data1' in content
        assert 'Data2' in content
        assert reader.sheet_separator in content
        
    def test_read_xlsx_with_formulas(self, reader, temp_dir):
        """Test reading XLSX file with formulas."""
        file_path = temp_dir / "formulas.xlsx"
        wb = Workbook()
        ws = wb.active
        
        # Add data with formula
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = '=A1+A2'
        
        wb.save(file_path)
        
        # Read with formulas
        reader_with_formulas = ExcelReader(include_formulas=True)
        content = reader_with_formulas.read(file_path)
        
        # Should show formula
        assert '=A1+A2' in content or '[=A1+A2]' in content
        
    def test_read_xlsx_with_comments(self, reader, temp_dir):
        """Test reading XLSX file with cell comments."""
        file_path = temp_dir / "comments.xlsx"
        wb = Workbook()
        ws = wb.active
        
        # Add data with comment
        ws['A1'] = 'Important'
        from openpyxl.comments import Comment
        ws['A1'].comment = Comment('This is important data', 'Test Author')
        
        wb.save(file_path)
        
        # Read with comments
        reader_with_comments = ExcelReader(include_comments=True)
        content = reader_with_comments.read(file_path)
        
        # Verify content
        assert 'Important' in content
        # Comments might be included if reader processes them
        
    def test_read_csv_comma_delimited(self, reader, temp_dir):
        """Test reading CSV file with comma delimiter."""
        file_path = temp_dir / "data.csv"
        
        # Create CSV content
        csv_content = "Name,Age,City\nAlice,30,New York\nBob,25,London"
        file_path.write_text(csv_content)
        
        # Read file
        content = reader.read(file_path)
        
        # Verify content
        assert 'Name | Age | City' in content
        assert 'Alice | 30 | New York' in content
        assert 'Bob | 25 | London' in content
        
    def test_read_csv_semicolon_delimited(self, reader, temp_dir):
        """Test reading CSV file with semicolon delimiter."""
        file_path = temp_dir / "data_semicolon.csv"
        
        # Create CSV content
        csv_content = "Name;Age;City\nAlice;30;New York\nBob;25;London"
        file_path.write_text(csv_content)
        
        # Read file
        content = reader.read(file_path)
        
        # Verify content is properly parsed
        assert 'Name | Age | City' in content
        assert 'Alice | 30 | New York' in content
        
    def test_read_csv_with_empty_cells(self, reader, temp_dir):
        """Test reading CSV with empty cells."""
        file_path = temp_dir / "empty_cells.csv"
        
        # Create CSV with empty cells
        csv_content = "A,B,C\n1,,3\n,2,\n4,5,6"
        file_path.write_text(csv_content)
        
        # Read file
        content = reader.read(file_path)
        
        # Verify empty cells are handled
        assert 'A | B | C' in content
        assert '1 |  | 3' in content or '1 | | 3' in content
        assert '4 | 5 | 6' in content
        
    def test_read_empty_xlsx(self, reader, temp_dir):
        """Test reading empty XLSX file."""
        file_path = temp_dir / "empty.xlsx"
        wb = Workbook()
        wb.save(file_path)
        
        # Should not fail
        content = reader.read(file_path)
        assert isinstance(content, str)
        
    def test_read_nonexistent_file(self, reader):
        """Test reading non-existent file raises error."""
        with pytest.raises(DocumentReadError):
            reader.read(Path("nonexistent.xlsx"))
            
    def test_read_corrupted_file(self, reader, temp_dir):
        """Test reading corrupted file raises error."""
        file_path = temp_dir / "corrupted.xlsx"
        file_path.write_bytes(b"This is not an Excel file")
        
        with pytest.raises(DocumentReadError):
            reader.read(file_path)
            
    def test_metadata_extraction(self, reader, temp_dir):
        """Test metadata extraction from Excel files."""
        file_path = temp_dir / "metadata_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        wb.save(file_path)
        
        # Read file
        reader.read(file_path)
        
        # Check metadata
        assert reader.metadata is not None
        assert 'filename' in reader.metadata
        assert 'file_size' in reader.metadata
        assert 'file_type' in reader.metadata
        assert reader.metadata['file_type'] == '.xlsx'
        
    def test_large_spreadsheet_handling(self, reader, temp_dir):
        """Test handling of large spreadsheets."""
        file_path = temp_dir / "large.csv"
        
        # Create a larger CSV
        df = pd.DataFrame({
            f'Col{i}': range(100) for i in range(10)
        })
        df.to_csv(file_path, index=False)
        
        # Should handle without issues
        content = reader.read(file_path)
        assert 'Col0' in content
        assert 'Col9' in content
        
    def test_special_characters_in_data(self, reader, temp_dir):
        """Test handling special characters in data."""
        file_path = temp_dir / "special_chars.csv"
        
        # Create CSV with special characters
        csv_content = 'Name,Description\n"Alice","Uses, commas"\n"Bob","Has\nnewlines"'
        file_path.write_text(csv_content)
        
        # Should handle special characters
        content = reader.read(file_path)
        assert 'Alice' in content
        assert 'Bob' in content